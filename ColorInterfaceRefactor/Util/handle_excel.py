'''
 * @Descripttion : 封装excel操作
 * @Author       : Tommy
 * @Date         : 2021-06-17 14:28:30
 * @LastEditors  : Tommy
 * @LastEditTime : 2021-07-19 18:27:45
'''
import openpyxl
from openpyxl.styles import Alignment


class HandleExcel(object):
    def __init__(self, filename):
        '''
         * @name: Tommy
         * @msg: HandleExcel初始化
         * @param {文件名/文件路径}
         * @return {}
        '''
        self.filename = filename
        self.excel_style = Alignment(horizontal='general',
                                     vertical='bottom',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

    def excel_open(self):
        '''
         * @name: Tommy
         * @msg: 打开excel文件
         * @param {}
         * @return {返回excel对象}
        '''
        excel_object = openpyxl.load_workbook(self.filename)
        return excel_object

    def get_table_by_index(self, index=None):
        '''
         * @name: Tommy
         * @msg: 获取表对象
         * @param {表索引号，默认为0}
         * @return {返回excel sheet object对象}
        '''
        excel_object = self.excel_open()
        sheet_name = excel_object.sheetnames
        if index is None:
            index = 0
        data = excel_object[sheet_name[index]]
        return data

    def get_cell_value(self, row, col, index=None):
        '''
         * @name: Tommy
         * @msg: excel单元格数据返回
         * @param {int: 行数, int: 列数, int: 表索引号}
         * @return {str：返回单元格数据}
        '''
        data = self.get_table_by_index(index).cell(row=row, column=col).value
        return data

    def get_rows(self, index=None):
        '''
         * @name: Tommy
         * @msg: 返回表的行数
         * @param {表索引号}
         * @return {int: 行数}
        '''
        row = self.get_table_by_index(index).max_row
        return row

    def get_rows_value(self, row, index=None):
        '''
         * @name: Tommy
         * @msg: 获取行数据
         * @param {int:行数, int:表索引号}
         * @return {list:返回对应row号的所有数据}
        '''
        row_list = []
        for i in self.get_table_by_index(index)[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        '''
         * @name: Tommy
         * @msg: 写入数据到excel
         * @param {row:写入行,cols:写入列,value:写入数据}
         * @return {}
        '''
        wb = self.excel_open()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(self.filename)

    def get_columns_values(self, cols, index=None):
        '''
         * @name: Tommy
         * @msg: 获取对应列的全部数据
         * @param {cols:列索引,index:表索引}
         * @return {返回数据集合}
        '''
        columns_list = []
        for i in self.get_table_by_index(index)[cols]:
            columns_list.append(i.value)
        return columns_list

    def get_row_number(self, data, key, index=None):
        '''
         * @name: Tommy
         * @msg: 根据接收数据获取对应行书
         * @param {data:接收数据, key:列关键字, index:表索引}
         * @return {返回行号/None}
        '''
        num = 0
        cols_data = self.get_columns_values(key, index)
        for col_data in cols_data:
            num += 1
            if data == col_data:
                return num

        return None

    def get_excel_data(self, index=None):
        '''
         * @name: Tommy
         * @msg: 返回所有excel数据按行获取
         * @param {index:表索引}
         * @return {返回所有数据，按行存为list，返回2维list}
        '''  
        data_list = []
        for i in range(self.get_rows(index)):
            data_list.append(self.get_rows_value(i + 2, index))

        return data_list


excel = HandleExcel(r"C:\Users\xt875\Documents\ToolSuit\ColorInterfaceRefactor\Case\Case.xlsx")

if __name__ == "__main__":
    handle = HandleExcel(
        r"C:\Users\xt875\Documents\ToolSuit\ColorInterfaceRefactor\Case\Case.xlsx"
    )
    value = handle.get_row_number('color_14', 'A')
    result = handle.get_cell_value(2, 10)
    print(value)
    print(result)
