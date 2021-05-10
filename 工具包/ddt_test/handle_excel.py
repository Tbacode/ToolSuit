'''
 * @Descripttion : 
 * @Author       : Tommy
 * @Date         : 2021-01-07 17:15:33
 * @LastEditors  : Tommy
 * @LastEditTime : 2021-04-21 17:49:20
'''
import openpyxl
from xlutils.copy import copy


class HandleExcel(object):
    def __init__(self, filename):
        self.filename = filename

    # TODO: 打开excel文件
    def excel_open(self):
        excel_object = openpyxl.load_workbook(self.filename)
        return excel_object

    # TODO: 获取表对象
    def get_table_by_index(self, index=None):
        sheet_name = self.excel_open().sheetnames
        if index is None:
            index = 0
        data = self.excel_open()[sheet_name[index]]
        return data

    # TODO: 返回表的行数
    def get_rows(self, index=None):
        row = self.get_table_by_index(index).max_row
        return row

    # TODO: 获取行数据
    def get_rows_value(self, row, index=None):
        row_list = []
        for i in self.get_table_by_index(index)[row]:
            row_list.append(i.value)
        return row_list

    # TODO: 获取所有数据
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows() - 1):
            data_list.append(self.get_rows_value(i + 2))
        return data_list

    # TODO: 返回表的列数
    def get_cols(self, index=None):
        col = self.get_table_by_index(index).max_column
        return col

    # TODO: excel单元格数据返回
    def get_cell_value(self, row, col, index=None):
        data = self.get_table_by_index(index).cell(row=row, column=col).value
        return data

    # TODO: excel追加数据
    def excel_celldata_add(self, row, col, sheet_name, content):
        # 打开表格
        excel_object = self.excel_open()
        # 将xlrd对象转换成xlwt对象
        xlwt_object = copy(excel_object)
        # 获取写入对象表
        table_object = xlwt_object.get_sheet(sheet_name)
        # 追加内容到单元格
        table_object.write(row, col, content)
        # 保存并覆盖
        xlwt_object.save(self.filename)