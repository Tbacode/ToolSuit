'''
 * @Descripttion : 
 * @Author       : Tommy
 * @Date         : 2021-01-07 17:15:33
 * @LastEditors  : Tommy
 * @LastEditTime : 2021-06-03 11:33:20
'''
import openpyxl
from openpyxl.styles import Alignment


class HandleExcel(object):
    def __init__(self, filename):
        '''
         * @name: Tommy
         * @msg: HandleExcel初始化
         * @param {文件名/文件路径}
         * @return {}
        '''
        self.filename = filename
        self.excel_style = Alignment(horizontal='general',
                                     vertical='bottom',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

    def excel_open(self):
        '''
         * @name: Tommy
         * @msg: 打开excel文件
         * @param {}
         * @return {返回excel object对象}
        '''
        excel_object = openpyxl.load_workbook(self.filename)
        return excel_object

    def get_table_by_index(self, index=None):
        '''
         * @name: Tommy
         * @msg: 获取表对象
         * @param {表索引号，默认为0}
         * @return {返回excel sheet object对象}
        '''
        sheet_name = self.excel_open().sheetnames
        if index is None:
            index = 0
        data = self.excel_open()[sheet_name[index]]
        return data

    def get_rows(self, index=None):
        '''
         * @name: Tommy
         * @msg: 返回表的行数
         * @param {表索引号}
         * @return {int: 行数}
        '''
        row = self.get_table_by_index(index).max_row
        return row

    def get_rows_value(self, row, index=None):
        '''
         * @name: Tommy
         * @msg: 获取行数据
         * @param {int:行数, int:表索引号}
         * @return {list:返回对应row号的所有数据}
        '''
        row_list = []
        for i in self.get_table_by_index(index)[row]:
            row_list.append(i.value)
        return row_list

    def get_excel_data(self):
        '''
         * @name: Tommy
         * @msg: 获取整个表的所有数据
         * @param {}
         * @return {list:返回整张表的所有数据}
        '''
        data_list = []
        for i in range(self.get_rows() - 1):
            data_list.append(self.get_rows_value(i + 2))
        return data_list

    def get_cols(self, index=None):
        '''
         * @name: Tommy
         * @msg: 返回表的列数
         * @param {int:表索引号}
         * @return {int:返回行数}
        '''
        col = self.get_table_by_index(index).max_column
        return col

    def get_cell_value(self, row, col, index=None):
        '''
         * @name: Tommy
         * @msg: excel单元格数据返回
         * @param {int: 行数, int: 列数, int:表索引号}
         * @return {str：返回单元格数据}
        '''
        data = self.get_table_by_index(index).cell(row=row, column=col).value
        return data

    def excel_celldata_add(self, row, col, sheet_name, content):
        '''
         * @name: Tommy
         * @msg: excel追加数据
         * @param {int: 行数, int: 列数, str: 表名, str: 追加内容}
         * @return {}
        '''
        # 打开表格
        excel_object = self.excel_open()
        # 创建一个新得sheet
        sheet = excel_object[sheet_name]
        # 写入数据
        sheet.cell(row=row, column=col,
                   value=content).alignment = self.excel_style
        # 保存
        excel_object.save(self.filename)